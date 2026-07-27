# ==========================================
# step3_export.py
# Xuất dữ liệu ra Excel
# ==========================================

import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import OUTPUT_EXCEL


def export_excel(data):

    print("\n" + "=" * 60)
    print("BƯỚC 3 - XUẤT EXCEL")
    print("=" * 60)

    if len(data) == 0:
        print("Không có dữ liệu để xuất.")
        return None

    # =====================================
    # DataFrame
    # =====================================

    df = pd.DataFrame(data)

    df.to_excel(
        OUTPUT_EXCEL,
        index=False
    )

    # =====================================
    # Mở workbook
    # =====================================

    wb = load_workbook(OUTPUT_EXCEL)

    ws = wb.active

    ws.title = "HoaDon"

    # =====================================
    # Style Header
    # =====================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin = Side(style="thin")

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # =====================================
    # Freeze
    # =====================================

    ws.freeze_panes = "A2"

    # =====================================
    # Filter
    # =====================================

    ws.auto_filter.ref = ws.dimensions

    # =====================================
    # Auto Width
    # =====================================

    for column_cells in ws.columns:

        length = 0

        letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:

                if cell.value is not None:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

            except:
                pass

        ws.column_dimensions[letter].width = min(length + 3, 50)

    # =====================================
    # Định dạng số
    # =====================================

    money_columns = [

        "Số lượng",
        "Đơn giá",
        "Chiết khấu",
        "Thành tiền",
        "Tiền thuế",
        "Tổng tiền trước thuế",
        "Tổng tiền thuế",
        "Tổng thanh toán"

    ]

    header = {}

    for cell in ws[1]:
        header[cell.value] = cell.column

    for col_name in money_columns:

        if col_name not in header:
            continue

        col = header[col_name]

        for row in range(2, ws.max_row + 1):

            ws.cell(row=row, column=col).number_format = '#,##0'

    # =====================================
    # Table
    # =====================================

    end_col = get_column_letter(ws.max_column)

    end_row = ws.max_row

    table = Table(
        displayName="InvoiceTable",
        ref=f"A1:{end_col}{end_row}"
    )

    style = TableStyleInfo(

        name="TableStyleMedium2",

        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False

    )

    table.tableStyleInfo = style

    ws.add_table(table)

    wb.save(OUTPUT_EXCEL)

    print(f"Đã xuất Excel:\n{OUTPUT_EXCEL}")

    print("=" * 60)

    return OUTPUT_EXCEL
